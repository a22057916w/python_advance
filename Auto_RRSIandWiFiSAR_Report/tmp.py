##! python3
##==============================================================================
##    Copyright (c) 2021 COMPAL Electronic Inc. All rights reserved.
##    This program contains proprietary and confidential information.
##    All rights reserved except as may be permitted by prior written consent.
##
##    Compal STiD NPSD Test Program Release Notification.
##
##    ModuleName:
##            report.py
##
##    Abstract:
##
##
##    Author:
##            22-Nov-2021 Willy Chen
##
##    Revision History:
##            Rev 1.0.0.1 22-Nov-2021 Willy
##                    First create.
##==============================================================================
import os, sys
import traceback
import shutil
import socket
import requests
import pandas as pd
import re
import time
import datetime as df
import codecs
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis


# [Main]
g_strVersion = "1.0.0.1"

#[Webside progress bar]
g_nProgressC = 0


def get_host_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()

    return ip

# Update UI Process (%)
def updateWebpageInfo(nProgress=g_nProgressC, strWebpageInfo=None):
    global g_nProgressC
    strWebComPath = os.path.join(logPath, "WebComunicate_%s.txt" % strUser)
    if nProgress != g_nProgressC:
        g_nProgressC = nProgress

    if nProgress == 0:
        ## Renew WebComunicate.txt
        with open(strWebComPath, 'w') as f:
            f.write('')
    else:
        strToWrite = "%s;%s" % (nProgress, strWebpageInfo)
        with open(strWebComPath, 'w') as f:
            f.write(strToWrite)




class Automation_RSSI_WiFiSARQUERY():

    strStartTime = "2021-11-26 11:14:20"
    strEdnTime = "2021-11-26 13:13:13"
    strProjectName = "./fuckme"

    def __init__(self, strUser="WillyWJ_Chen", b_localDebug=True):
        self.setPath(strUser, b_localDebug)

    def start(self):
        # get directory names of TryingLog
        self.listSNLogs = os.listdir(self.inputFolder)
        # iterate through log files in a SN folder and get parsed data
        self.listRSSI, self.listWIFI = parseLog(self.listSNLogs, self.inputFolder)
        # save parsed data to excel
        log_to_excel(self.listRSSI, self.listWIFI, self.outputFolder)

    def setPath(self, strUser, b_localDebug):
        if not b_localDebug:
            # Server upload path
            SourcePath = "/home/sanchez/Desktop/webserver/ToolPage/server/php/files/ToolPage/Automation_RSSI_WiFiSARQUERY/%s" % strUser
            # InputFolder
            InputFolder = "/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY/input_%s" % strUser
            # Output file in download folder
            ResultPath = "/home/sanchez/Desktop/webserver/ToolPage/Download/Automation_RSSI_WiFiSARQUERY_%s.xlsx" % strUser
            # Output Path
            outputFolder = "/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY/output_%s" % strUser
        else:
            self.dataPath = "./test_SN"                     # raw data folder(source to be parsed)
            self.mappingJsonPath = ".\mapping\sanchezPeng\mapping.json"
            self.inputFolder = "./input/%s" % strUser       # Output file in download folder
            self.outputFolder = "./output/%s" % strUser     # Output Path

        list_path = [self.inputFolder, self.outputFolder]
        for path in list_path:
            self.initPath(path)

        self.mapping(self.mappingJsonPath)
        self.pullData()

    def initPath(self, strDirPath):
        try:
            if os.path.exists(strDirPath):
                shutil.rmtree(strDirPath)
                printLog("[I][initPath] Delete Folder: %s" % strDirPath)

            os.mkdir(strDirPath)
            printLog("[I][initPath] Create Folder: %s" % strDirPath)
        except Exception as e:
            printLog("[E][initPaht] Unexpected Error: " + str(e))
            print("[E][initPath] Unexception error: %s" % str(sys.exc_info()[1]))

    def mapping(self, strJsonPath):
        try:
            with open(strJsonPath, "r") as jsonFile:
                dictInfo = json.load(jsonFile)
                self.strStartTime = dictInfo["starttime"]
                self.strEdnTime = dictInfo["endtime"]
                self.strProjectName = dictInfo["Projectname"]
        except Exception as e:
            printLog("[E][mapping] Unexpected Error: " + str(e))
            print("[E][mapping] Unexception error: %s" % str(sys.exc_info()[1]))

    def pullData(self):
        try:
            start_time = df.datetime.strptime(self.strStartTime, "%Y-%m-%d %H:%M:%S")
            end_time = df.datetime.strptime(self.strEdnTime, "%Y-%m-%d %H:%M:%S")

            #start_time = df.datetime.strptime("2021-11-26 11:14:20", "%Y-%m-%d %H:%M:%S")
            #end_time = df.datetime.strptime("2021-11-26 13:13:13", "%Y-%m-%d %H:%M:%S")

            for SN_dir in os.listdir(self.dataPath):
                SN_path = os.path.join(self.dataPath, SN_dir)

                c_epoch_time = os.path.getctime(SN_path)
                c_time = df.datetime.fromtimestamp(c_epoch_time)

                if c_time >= start_time and c_time <= end_time:
                    shutil.copytree(SN_path, os.path.join(self.inputFolder, SN_dir))
                    print(SN_dir)
        except Exception as e:
            printLog("[E][pullData] Unexpected error: " + str(e))
            print("[E][pullData] Unexception error: %s" % str(sys.exc_info()[1]))


#/====================================================================\#
#|               Functions of parsing target logs                     |#
#\====================================================================/#

# retunn two list of dict, listRSSI, listWIFI
def parseLog(list_SNs, strInuptFolder):
    printLog("[I][parseLog] ------- Start Parsing Log -------")

    listRSSI, listWIFI = [], []
    try:
        for strSN in list_SNs:
            strRSSI_Intel, strRSSI_MTK = "ITPM_RSSITest_A32.LOG", "MTKlog_" + strSN + ".LOG"  # e.g. MTKlog_7228979800008.LOG
            strWIFI = "SarLog_DynAnt.txt"

            # for RSSI excel columns and WIFI's
            dictRSSI = {
                "UNIT PN" : strSN,
                "Test Result" : None,
                "Main" : None,
                "Aux" : None,
                "Spec" : None }
            dictWIFI = {
                "UNIT PN" : strSN,
                "Result" : "no log" }

            b_hasRSSI, b_hasWIFI = False, False             # flag for checking if the target log exists
            strSNPath = os.path.join(strInuptFolder, strSN)    # set abspath for SN logs

            # iterate through log files in a SN folder
            for strFileName in os.listdir(strSNPath):
                strLogPath = os.path.join(strSNPath, strFileName)

                # check if ITPM_RSSITest_A32.log exists. If not, flag = False and parse only SN.
                if strFileName == strRSSI_Intel:
                    parseRSSI_Intel(dictRSSI, strLogPath)
                    b_hasRSSI = True

                # if MTKlog_7228979800008.LOG exists, parse the log
                if strFileName == strRSSI_MTK:
                    parseRSSI_MTK(dictRSSI, strLogPath)
                    b_hasRSSI = True

                # parse SarLog_DynAnt.txt
                if strFileName == strWIFI:
                    parseWIFI(dictWIFI, strLogPath)
                    b_hasWIFI = True

            # if log not exists, append initial dict
            listRSSI.append(dictRSSI)
            listWIFI.append(dictWIFI)

            # if there is no target log file in the folder, parse only SN
            if not b_hasRSSI:
                printLog("[W][ParseLog] Cannot find %s and %s by SN: %s" % (strRSSI_Intel, strRSSI_MTK, strSN))
            if not b_hasWIFI:
                printLog("[W][ParseLog] Cannot find %s by SN: %s" % (strWIFI, strSN))

        printLog("[I][parseLog] ------- Finish Parsing Log -------")
    except Exception as e:
        printLog("[E][parseLog] Unexpected Error: " + str(e))
    return listRSSI, listWIFI

# parse line contains 'CODE NO="010001A0"' and 'CODE NO="01000120"'
def parseRSSI_Intel(dictRSSI, strRSSIPath):
    try:
        with open(strRSSIPath, encoding="big5") as RSSILog:    # big5 for windows
            content = RSSILog.readlines()
            for line in content:
                # if the first line is PASS or FAIL, save to dict
                if content.index(line) == 0 and ("PASS" in line or "FAIL" in line):
                    dictRSSI["Test Result"] = line.strip("\n")

                if 'CODE NO="010001A0"' in line:
                    # line return <CODE NO="010001A0" KEYNAME="RSSI" VALUE1="-65" VALUE2="-63" ...
                    list_Info = line.split(" ")

                    # get VALUE1 figure
                    strSpec = list_Info[3].split("=")[1].strip('/>"')
                    dictRSSI["Spec"] = eval(strSpec)

                    # get VALUE2 figure
                    strMain = list_Info[4].split("=")[1].strip('/>"')
                    dictRSSI["Main"] = eval(strMain)

                if 'CODE NO="01000120"' in line:
                    # line return <CODE NO="01000120" KEYNAME="RSSI" VALUE1="-65" VALUE2="-62" ...
                    list_Info = line.split(" ")

                    strAux = list_Info[4].split("=")[1].strip('/>"')
                    dictRSSI["Aux"] = eval(strAux)
    except Exception as e:
        printLog("[E][parseRSSI_Intel] Unexpected Error: " + str(e))

# parse line contains "== ANT1 (sub) ==" and "== ANT2 (main) =="
def parseRSSI_MTK(dictRSSI, strRSSIPath):
    try:
        with open(strRSSIPath, encoding="big5") as RSSILog:    # big5 for windows
            content = RSSILog.readlines()
            for line in content:

                if "== ANT1 (sub) ==" in line:
                    idx = content.index(line)       # get the line index of keyword
                    for line in content[idx:]:      # start iteration from the line "== ANT1 (sub) =="
                        if "Threshold" in line:
                            strThreshold = line.split(": ")[1].strip("\n")
                            dictRSSI["Spec"] = eval(strThreshold)
                        if "Average" in line:
                            strMain = line.split(": ")[1].strip("\n")
                            dictRSSI["Main"] = eval(strMain)
                            break                   # find only one keyword "Average"

                if "== ANT2 (main) ==" in line:
                    idx = content.index(line)       # get the line index of keyword
                    for line in content[idx:]:
                        if "Average" in line:
                            strAux = line.split(": ")[1].strip("\n")
                            dictRSSI["Aux"] = eval(strAux)
                        if "PASS" in line or "FAIL" in line:
                            strResult = line.strip("\n")
                            dictRSSI["Test Result"] = strResult
                            break
    except Exception as e:
        print("[E][parseRSSI_MTK] Unexpected Error: " + str(e))

# parse the last line in SarLog_DynAnt.txt, target = "Value Match" or "Value Not Match"
def parseWIFI(dictWIFI, strWIFIPath):
    try:
        with open(strWIFIPath, encoding="big5") as WIFILog:    # big5 for windows
            content = WIFILog.readlines()
            # the target is in the last line
            if "Value Match" in content[-1] or "Value Not Match" in content[-1]:
                strResult = content[-1].strip("\n")
                dictWIFI["Result"] = strResult
    except Exception as e:
        printLog("[E][parseWIFI] Unexpected Error: " + str(e))

#/====================================================================\#
#|              Functions of parsing log to excel                     |#
#\====================================================================/#

# save two xlsx, RSSI.xlsx, WiFiSARQUERY.xlsx
def log_to_excel(listRSSI, listWIFI, strOutputFolder):
    printLog("[I][log_to_excel] ------- Parsing Log to Excel -------")

    try:
        # covert list of dict to DataFrame in order
        df_logRSSI = pd.DataFrame(listRSSI, columns=listRSSI[0].keys())
        df_logWIFI = pd.DataFrame(listWIFI, columns=listWIFI[0].keys())

        # set df_logRSSI index
        df_logRSSI.index.name = "Item"
        df_logRSSI.index += 1                   # index start with 1
        df_logRSSI = df_logRSSI.reset_index()   # reset index as normal column for constructing sheet

        # list the arguments for createing Excel workbooks
        list_df = [df_logRSSI, df_logWIFI]
        list_sheetname = ["RSSI", "wifisarquery"]
        list_fname = ["RSSI.xlsx", "WiFiSARQUERY.xlsx"]

        # new Excel workbook and sheets
        for i in range(len(list_df)):
            wb = openpyxl.Workbook()    # 新增 Excel 活頁
            wb.remove(wb['Sheet'])      # remove the default sheet when start a workbook

            # init vars
            df = list_df[i]
            str_sheet = list_sheetname[i]
            str_fname = list_fname[i]

            # set up sheet by DataFrame
            newSheet(wb, str_sheet, df)
            styleSheet(wb[str_sheet])
            # plot for RSSI_Report.xlsx
            if i == 0:
                newLineChart(wb[str_sheet], df)
                #print(df)
            wb.save(os.path.join(strOutputFolder, str_fname))     # save the worksheet as excel file
            printLog("[I][log_to_excel] %s generated" % str_fname)

        printLog("[I][log_to_excel] ------- Parsed Log to Excel -------")
    except Exception as e:
        printLog("[E][log_to_excel] Unexpected Error: " + str(e))

# new worksheets by DataFrame
def newSheet(workbook, strSheetName, df):
    try:
        workbook.create_sheet(strSheetName)
        for row in dataframe_to_rows(df, index=False, header=True):
            workbook[strSheetName].append(row)

        printLog("[I][newSheet] Sheet: %s Created" % strSheetName)
    except Exception as e:
        printLog("[E][newSheet] Unexpected Error: " + str(e))

def styleSheet(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))

    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value + 3
    dims.clear()

    # Enumerate the cells in the second row
    for cell in ws["1:1"]:
        cell.font = Font(color="F8F8FF")    # ghostwhite
        cell.fill = PatternFill(fgColor="292421", fill_type="solid")

# draw a line chart with 3 lines for RSSI
def newLineChart(ws, df):
    try:
        chart = LineChart()
        chart.title = "UNIT PN_M"
        chart.style = 18

        data = Reference(ws, min_col=4, min_row=1, max_col=6, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)

        # set daigram size
        chart.height = 10   # default is 7.5
        chart.width = 20

        # set y-axis according to the max value from DataFrame(df_logRSSI)
        chart.y_axis.scaling.max = -30

        # style the lines
        line_Main = chart.series[0]
        line_Main.graphicalProperties.line.solidFill = "00AAAA"     # navyblue
        line_Main.graphicalProperties.line.width = 25000

        line_Aux = chart.series[1]
        line_Aux.graphicalProperties.line.solidFill = "F08080"      # lightpink
        line_Aux.graphicalProperties.line.width = 25000

        line_Spec = chart.series[2]
        line_Spec.graphicalProperties.line.solidFill = "3D9140"     # cobaltgreen
        line_Spec.graphicalProperties.line.width = 25000

        ws.add_chart(chart, "I1")
    except Exception as e:
        printLog("[E][newLineChart] Unexpected Error: " + str(e))


#/====================================================================\#
#|                        Functions of printing log                   |#
#\====================================================================/#

def getDateTimeFormat():
    strDateTime = "[%s]" % (time.strftime("%Y/%m/%d %H:%M:%S", time.localtime()))
    return strDateTime

def printLog(strPrintLine):
    global strUser
    strFileName = os.path.basename(__file__).split('.')[0]
    fileLog = codecs.open(strFileName + "_" + strUser + ".log", 'a', "utf-8")
    print(strPrintLine)
    fileLog.write("%s%s\r\n" % (getDateTimeFormat(), strPrintLine))
    fileLog.close()



def zip_all_files(self):
        try:
            self.module_logger.info("------------ Start Zipping Files. ------------")
            with zipfile.ZipFile(os.path.join(self.str_output_path, "MultiLevel_%s.zip"%strUser), "w") as zf:
                for dirPath, dirNames, fileNames in os.walk(self.str_output_path):
                    for f in fileNames:
                        if "zip" not in f:
                            #print(88888,os.path.join(dirPath, f))
                            zf.write(os.path.join(dirPath, f), os.path.basename(f))
            updateWebpageInfo(97, "------------ Zip All Files Done. ------------")
            return True

        except:
            self.module_logger.error("Unexpected error: %s" % (str(traceback.format_exc())))
            print("[E][zip_all_files] Exception error: %s" % str(sys.exc_info()[1]))
            return False

if __name__ == "__main__":
    global strUser

    if(len(sys.argv) < 1):
        strUser = "WillyWJ_Chen"
    else:
        strUser = sys.argv[1]

    strFileName = os.path.basename(__file__).split('.')[0]

    printLog("========== Start ==========")
    printLog("[I][main] Python " + sys.version)
    printLog("[I][main] %s.py %s" % (strFileName, g_strVersion))


    logPath = "/home/sanchez/Desktop/RDTool/CTO_R1R3_Checking"

    try:
        ARW = Automation_RSSI_WiFiSARQUERY()
        ARW.start()

    except Exception as e:
        printLog("[E][main] Unexpected Error: " + str(e))
    printLog("========== End ==========")

##! python3
##==============================================================================
##    Copyright (c) 2021 COMPAL Electronic Inc. All rights reserved.
##    This program contains proprietary and confidential information.
##    All rights reserved except as may be permitted by prior written consent.
##
##    Compal STiD NPSD Test Program Release Notification.
##
##    ModuleName:
##            report.py
##
##    Abstract:
##
##
##    Author:
##            22-Nov-2021 Willy Chen
##
##    Revision History:
##            Rev 1.0.0.1 22-Nov-2021 Willy
##                    First create.
##==============================================================================
import os, sys
import traceback
import shutil
import socket
import requests
import pandas as pd
import re
import time
import codecs
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis


# [Main]
g_strVersion = "1.0.0.1"

#[ParseLogPath]
g_strLogDir = "./test_SN/"

#[Webside progress bar]
g_nProgressC = 0


def get_host_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()

    return ip


# Update UI Process (%)
def updateWebpageInfo(nProgress=g_nProgressC, strWebpageInfo=None):
    global g_nProgressC
    strWebComPath = os.path.join(logPath, "WebComunicate_%s.txt" % strUser)
    if nProgress != g_nProgressC:
        g_nProgressC = nProgress

    if nProgress == 0:
        ## Renew WebComunicate.txt
        with open(strWebComPath, 'w') as f:
            f.write('')
    else:
        strToWrite = "%s;%s" % (nProgress, strWebpageInfo)
        with open(strWebComPath, 'w') as f:
            f.write(strToWrite)


#/====================================================================\#
#|               Functions of parsing target logs                     |#
#\====================================================================/#

def parseLog(list_SNs):
    printLog("[I][parseLog] ------- Start Parsing Log -------")

    listRSSI, listWIFI = [], []
    try:
        for strSN in list_SNs:
            strRSSI_Intel, strRSSI_MTK = "ITPM_RSSITest_A32.LOG", "MTKlog_" + strSN + ".LOG"  # e.g. MTKlog_7228979800008.LOG
            strWIFI = "SarLog_DynAnt.txt"

            # for RSSI excel columns and WIFI's
            dictRSSI = {
                "UNIT PN" : strSN,
                "Test Result" : None,
                "Main" : None,
                "Aux" : None,
                "Spec" : None }
            dictWIFI = {
                "UNIT PN" : strSN,
                "Result" : "no log" }

            b_hasRSSI, b_hasWIFI = False, False             # flag for checking if the target log exists
            strSNPath = os.path.join(g_strLogDir, strSN)    # set abspath for SN logs

            # iterate through log files in a SN folder
            for strFileName in os.listdir(strSNPath):
                strLogPath = os.path.join(strSNPath, strFileName)

                # check if ITPM_RSSITest_A32.log exists. If not, flag = False and parse only SN.
                if strFileName == strRSSI_Intel:
                    parseRSSI_Intel(dictRSSI, strLogPath)
                    b_hasRSSI = True

                # if MTKlog_7228979800008.LOG exists, parse the log
                if strFileName == strRSSI_MTK:
                    parseRSSI_MTK(dictRSSI, strLogPath)
                    b_hasRSSI = True

                # parse SarLog_DynAnt.txt
                if strFileName == strWIFI:
                    parseWIFI(dictWIFI, strLogPath)
                    b_hasWIFI = True

            # if log not exists, append initial dict
            listRSSI.append(dictRSSI)
            listWIFI.append(dictWIFI)

            # if there is no target log file in the folder, parse only SN
            if not b_hasRSSI:
                printLog("[W][ParseLog] Cannot find %s and %s by SN: %s" % (strRSSI_Intel, strRSSI_MTK, strSN))
            if not b_hasWIFI:
                printLog("[W][ParseLog] Cannot find %s by SN: %s" % (strWIFI, strSN))

        printLog("[I][parseLog] ------- Finish Parsing Log -------")
    except Exception as e:
        printLog("[E][parseLog] Unexpected Error: " + str(e))
    return listRSSI, listWIFI


def parseRSSI_Intel(dictRSSI, strRSSIPath):
    try:
        with open(strRSSIPath, encoding="big5") as RSSILog:    # big5 for windows
            content = RSSILog.readlines()
            for line in content:
                # if the first line is PASS or FAIL, save to dict
                if content.index(line) == 0 and ("PASS" in line or "FAIL" in line):
                    dictRSSI["Test Result"] = line.strip("\n")

                if 'CODE NO="010001A0"' in line:
                    # line return <CODE NO="010001A0" KEYNAME="RSSI" VALUE1="-65" VALUE2="-63" ...
                    list_Info = line.split(" ")

                    # get VALUE1 figure
                    strSpec = list_Info[3].split("=")[1].strip('/>"')
                    dictRSSI["Spec"] = eval(strSpec)

                    # get VALUE2 figure
                    strMain = list_Info[4].split("=")[1].strip('/>"')
                    dictRSSI["Main"] = eval(strMain)

                if 'CODE NO="01000120"' in line:
                    # line return <CODE NO="01000120" KEYNAME="RSSI" VALUE1="-65" VALUE2="-62" ...
                    list_Info = line.split(" ")

                    strAux = list_Info[4].split("=")[1].strip('/>"')
                    dictRSSI["Aux"] = eval(strAux)
    except Exception as e:
        printLog("[E][parseRSSI_Intel] Unexpected Error: " + str(e))


def parseRSSI_MTK(dictRSSI, strRSSIPath):
    try:
        with open(strRSSIPath, encoding="big5") as RSSILog:    # big5 for windows
            content = RSSILog.readlines()
            for line in content:

                if "== ANT1 (sub) ==" in line:
                    idx = content.index(line)       # get the line index of keyword
                    for line in content[idx:]:
                        if "Threshold" in line:
                            strThreshold = line.split(": ")[1].strip("\n")
                            dictRSSI["Spec"] = eval(strThreshold)
                        if "Average" in line:
                            strMain = line.split(": ")[1].strip("\n")
                            dictRSSI["Main"] = eval(strMain)
                            break

                if "== ANT2 (main) ==" in line:
                    idx = content.index(line)       # get the line index of keyword
                    for line in content[idx:]:
                        if "Average" in line:
                            strAux = line.split(": ")[1].strip("\n")
                            dictRSSI["Aux"] = eval(strAux)
                        if "PASS" in line or "FAIL" in line:
                            strResult = line.strip("\n")
                            dictRSSI["Test Result"] = strResult
                            break
    except Exception as e:
        print("[E][parseRSSI_MTK] Unexpected Error: " + str(e))


def parseWIFI(dictWIFI, strWIFIPath):
    try:
        with open(strWIFIPath, encoding="big5") as WIFILog:    # big5 for windows
            content = WIFILog.readlines()
            if "Value Match" in content[-1] or "Value Not Match" in content[-1]:
                strResult = content[-1].strip("\n")
                dictWIFI["Result"] = strResult
    except Exception as e:
        printLog("[E][parseWIFI] Unexpected Error: " + str(e))

#/====================================================================\#
#|              Functions of parsing log to excel                     |#
#\====================================================================/#

def log_to_excel(listRSSI, listWIFI):
    printLog("[I][log_to_excel] ------- Parsing Log to Excel -------")

    # ========== New Excel workbook and sheets ==========
    try:

        df_logRSSI = pd.DataFrame(listRSSI)
        df_logWIFI = pd.DataFrame(listWIFI)

        list_df = [df_logRSSI, df_logWIFI]
        list_sheetname = ["RSSI", "wifisarquery"]
        list_fname = ["RSSI_Report.xlsx", "WIFI_Report.xlsx"]


        for i in range(len(list_df)):
            wb = openpyxl.Workbook()    # 新增 Excel 活頁
            wb.remove(wb['Sheet'])      # remove the default sheet when start a workbook

            df = list_df[i]
            str_sheet = list_sheetname[i]
            # set up sheet by DataFrame
            newSheet(wb, str_sheet, df)

            if i == 0:
                newLineChart(wb[str_sheet], df)

            str_fname = list_fname[i]
            wb.save(str_fname)     # save the worksheet as excel file

        printLog("[I][log_to_excel] ------- Parsed Log to Excel -------")
    except Exception as e:
        printLog("[E][log_to_excel] Unexpected Error: " + str(e))

# new worksheets by DataFrame
def newSheet(workbook, strSheetName, df_SheetCol):
    try:
        workbook.create_sheet(strSheetName)
        for row in dataframe_to_rows(df_SheetCol, index=False, header=True):
            workbook[strSheetName].append(row)
        printLog("[I][newSheet] Sheet: %s Created" % strSheetName)
    except Exception as e:
        printLog("[E][newSheet] Unexpected Error: " + str(e))


def newLineChart(ws, df):
    c1 = LineChart()
    c1.title = "UNIT PN_M"
    c1.style = 13
    #c1.y_axis.title = 'Size'
    #c1.x_axis.title = 'Test Number'

    data = Reference(ws, min_col=3, min_row=1, max_col=5, max_row=7)
    c1.add_data(data, titles_from_data=True)

    #c1.y_axis.scaling.min = -100
    #c1.y_axis.scaling.max = -40
    line_Main = c1.series[0]
    line_Main.graphicalProperties.line.solidFill = "00AAAA"     # navyblue

    line_Aux = c1.series[1]
    line_Aux.graphicalProperties.line.solidFill = "F08080"      # lightpink

    line_Spec = c1.series[2]
    line_Spec.graphicalProperties.line.solidFill = "3D9140"     # cobaltgreen


    ws.add_chart(c1, "G1")


#/====================================================================\#
#|                        Functions of printing log                   |#
#\====================================================================/#

def getDateTimeFormat():
    strDateTime = "[%s]" % (time.strftime("%Y/%m/%d %H:%M:%S", time.localtime()))
    return strDateTime

def printLog(strPrintLine):
    strFileName = os.path.basename(__file__).split('.')[0]
    fileLog = codecs.open(g_strFileName + ".log", 'a', "utf-8")
    print(strPrintLine)
    fileLog.write("%s%s\r\n" % (getDateTimeFormat(), strPrintLine))
    fileLog.close()


if __name__ == "__main__":
    global g_strFileName, g_ShareFolder_ip
    g_strFileName = os.path.basename(__file__).split('.')[0]

    printLog("========== Start ==========")
    printLog("[I][main] Python " + sys.version)
    printLog("[I][main] %s.py %s" % (g_strFileName, g_strVersion))

    # ------------ find the target file --------------
    try:
        # get directory names of TryingLog (first layer)
        listSNLogs = os.listdir(g_strLogDir)
        # iterate through log files in a SN folder and get parsed data
        listRSSI, listWIFI = parseLog(listSNLogs)

        log_to_excel(listRSSI, listWIFI)


    except Exception as e:
        printLog("[E][main] Unexpected Error: " + str(e))
    printLog("========== End ==========")

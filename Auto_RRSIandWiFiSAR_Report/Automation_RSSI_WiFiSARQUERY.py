##! python3
##==============================================================================
##    Copyright (c) 2021 COMPAL Electronic Inc. All rights reserved.
##    This program contains proprietary and confidential information.
##    All rights reserved except as may be permitted by prior written consent.
##
##    Compal STiD NPSD Test Program Release Notification.
##
##    ModuleName:
##            Automation_RSSI_WiFiSARQUERY.py
##
##    Abstract:
##            1. Pulling data from share folder
##            2. Parsing log(data) into two xlsx-s
##            3. Saving result as zip file, then push to server
##    Author:
##            22-Nov-2021 Willy Chen
##
##    Revision History:
##            Rev 1.0.0.1 22-Nov-2021 Willy
##                    First create.
##==============================================================================
import os, sys
import traceback
import shutil
import socket
import requests
import pandas as pd
import re
import time
import datetime as df
import codecs
import json
import zipfile
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

# set share folder path
from pathlib import Path

# [Main]
g_strVersion = "1.0.0.1"

#[Webside progress bar]
g_nProgressC = 0


def get_host_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()

    return ip

# Update UI Process (%)
def updateWebpageInfo(nProgress=g_nProgressC, strWebpageInfo=None):
    global g_nProgressC
    strWebComPath = os.path.join(logPath, "WebComunicate_%s.txt" % strUser)
    if nProgress != g_nProgressC:
        g_nProgressC = nProgress

    if nProgress == 0:
        ## Renew WebComunicate.txt
        with open(strWebComPath, 'w') as f:
            f.write('')
    else:
        strToWrite = "%s;%s" % (nProgress, strWebpageInfo)
        with open(strWebComPath, 'w') as f:
            f.write(strToWrite)


#/====================================================================\#
#|                               Class                                |#
#\====================================================================/#

class Automation_RSSI_WiFiSARQUERY():

    # default value
    strStartTime = "2021-11-21 11:14:20"
    strEdnTime = "2021-11-26 13:13:13"
    strProjectName = "./fuckme"

    def __init__(self, strUser="WillyWJ_Chen", b_localDebug=True):
        self.setPath(strUser, b_localDebug)

    def start(self):
        printLog("\n[I][start] ------ Begin Generating Sequence -----")
        updateWebpageInfo(50, "------------ Parsing Log ------------")
        time.sleep(1)
        # get directory names of TryingLog
        self.listSNLogs = os.listdir(self.inputFolder)
        # iterate through log files in a SN folder and get parsed data
        self.listRSSI, self.listWIFI = parseLog(self.listSNLogs, self.inpuFolder)

        # save parsed data as excel xlsx
        updateWebpageInfo(75, "------------ Generating Excel files ------------")
        time.sleep(1)
        log_to_excel(self.listRSSI, self.listWIFI, self.outputFolder)

        # zip path for compressing two xlsx
        self.zipfilePath = os.path.join(self.outputFolder, "Automation_RRSIandWiFiSAR_Report_%s.zip" % strUser)

        # compressing two xlsx
        updateWebpageInfo(90, "------------ Compressing Excel files ------------")
        time.sleep(1)
        zip_all_files(self.outputFolder, self.zipfilePath)

        # copy result to server dir
        printLog("[I][start] Saving Zip File to Server")
        updateWebpageInfo(95, "------------ Saving Zip File to Server ------------")
        time.sleep(1)
        shutil.copy2(self.zipfilePath, self.resultPath)

        printLog("[I][start] ------ End of Generating Sequence -----")
        updateWebpageInfo(100, "------------ Finish ------------")

    # set folder path with strUser
    def setPath(self, strUser, b_localDebug):
        try:
            printLog("\n[I][setPath] ----- Setting Folder Path -----")
            updateWebpageInfo(5, "------------ Setting Path ------------")

            if not b_localDebug:
                # mapping criteria for pulling data
                self.mappingJsonPath = "/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY/Mapping/%s/Mapping.json" % strUser
                # InputFolder
                self.inputFolder = "/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY/input/%s" % strUser
                # Output Path
                self.outputFolder = "/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY/output/%s" % strUser
                # Output file in download folder
                self.resultPath = "/home/sanchez/Desktop/webserver/ToolPage/Download"
            else:
                self.mappingJsonPath = "./Mapping/SanchezPeng/Mapping.json"     # raw data folder(source to be parsed)
                self.inputFolder = "./input/%s" % strUser       # Output file in download folder
                self.outputFolder = "./output/%s" % strUser     # Output Path
                self.resultPath = os.getcwd()

            updateWebpageInfo(10, "------------ Setting Path ------------")
            time.sleep(5)

            # remove old files and new folders
            list_path = [self.inputFolder, self.outputFolder]
            for path in list_path:
                self.initPath(path)

            # mapping default variables, then pulling data as params
            updateWebpageInfo(12, "------------ Mapping Criteria ------------")
            time.sleep(1)
            self.mapping(self.mappingJsonPath)

            # set data path by the params from mapping
            #self.dataPath = '/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY/ARW_temp/%s' % (self.strProjectName)
            self.dataPath = './ARW_temp/%s' % (self.strProjectName)

            # pull data from self.dataPath, via share folder or not
            updateWebpageInfo(15, "------------ Pulling Data ------------")
            time.sleep(1)
            self.pullData()

            printLog("[I][setPath] ----- Setting Folder Path Successfully -----")
        except Exception as e:
            printLog("[E][setPath] Unexpected Error: " + str(e))
            print("[E][setPath] Unexception error: %s" % str(sys.exc_info()[1]))

    # remove old fils and create new folder
    def initPath(self, strDirPath):
        try:
            if os.path.exists(strDirPath):
                shutil.rmtree(strDirPath)
                printLog("[I][initPath] Delete Folder: %s" % strDirPath)

            os.mkdir(strDirPath)

            printLog("[I][initPath] Create Folder: %s" % strDirPath)
        except Exception as e:
            printLog("[E][initPaht] Unexpected Error: " + str(e))
            print("[E][initPath] Unexception error: %s" % str(sys.exc_info()[1]))

    # parse json data from ./Mapping/...
    def mapping(self, strJsonPath):
        try:
            with open(strJsonPath, "r") as jsonFile:
                dictInfo = json.load(jsonFile)
                self.strStartTime = dictInfo["starttime"]
                self.strEdnTime = dictInfo["endtime"]
                self.strProjectName = dictInfo["Projectname"]
        except Exception as e:
            printLog("[E][mapping] Unexpected Error: " + str(e))
            print("[E][mapping] Unexception error: %s" % str(sys.exc_info()[1]))

    # pull data from share folder
    def pullData(self):
        try:
            printLog("[I][pullData] ----- Pulling Data ------")

            start_time = df.datetime.strptime(self.strStartTime, "%Y-%m-%d %H:%M:%S")
            end_time = df.datetime.strptime(self.strEdnTime, "%Y-%m-%d %H:%M:%S")

            n_fCount = 0

            # copying files according to the creation date wihtin [start_time, end_time]
            for SN_dir in os.listdir(self.dataPath):
                SN_path = os.path.join(self.dataPath, SN_dir)

                c_epoch_time = os.path.getctime(SN_path)            # return the epoch time(float)
                c_time = df.datetime.fromtimestamp(c_epoch_time)    # convert the epoch time to human readible date

                if c_time >= start_time and c_time <= end_time:
                    shutil.copytree(SN_path, os.path.join(self.inputFolder, SN_dir))
                    n_fCount += 1
                    updateWebpageInfo(15 + n_fCount/len(os.listdir(self.dataPath)) * 35, "------------ Pulling Data ------------")
                    time.sleep(0.02)
                    #print(SN_dir)

            # check if there is at least one file pull
            if len(os.listdir(self.inputFolder)) > 0:
                printLog("[I][pullData] ----- Pulling Data Successfully -----")
            else:
                printLog("[W][pullData] No files wihtin [%s, %s]" % (start_time, end_time))
                sys.exit("Error: No files wihtin [%s, %s]" % (start_time, end_time))

        except Exception as e:
            printLog("[E][pullData] Unexpected error: " + str(e))
            print("[E][pullData] Unexception error: %s" % str(sys.exc_info()[1]))


#/====================================================================\#
#|               Functions of parsing target logs                     |#
#\====================================================================/#

# retunn two list of dict, listRSSI, listWIFI
def parseLog(list_SNs, strInuptFolder):
    printLog("[I][parseLog] ------- Start Parsing Log -------")

    listRSSI, listWIFI = [], []
    try:
        for strSN in list_SNs:
            strRSSI_Intel, strRSSI_MTK = "ITPM_RSSITest_A32.LOG", "MTKlog_" + strSN + ".LOG"  # e.g. MTKlog_7228979800008.LOG
            strWIFI = "SarLog_DynAnt.txt"

            # for RSSI excel columns and WIFI's
            dictRSSI = {
                "UNIT PN" : strSN,
                "Test Result" : None,
                "Main" : None,
                "Aux" : None,
                "Spec" : None }
            dictWIFI = {
                "UNIT PN" : strSN,
                "Result" : "no log" }

            b_hasRSSI, b_hasWIFI = False, False             # flag for checking if the target log exists
            strSNPath = os.path.join(strInuptFolder, strSN)    # set abspath for SN logs

            # iterate through log files in a SN folder
            for strFileName in os.listdir(strSNPath):
                strLogPath = os.path.join(strSNPath, strFileName)

                # check if ITPM_RSSITest_A32.log exists. If not, flag = False and parse only SN.
                if strFileName == strRSSI_Intel:
                    parseRSSI_Intel(dictRSSI, strLogPath)
                    b_hasRSSI = True

                # if MTKlog_7228979800008.LOG exists, parse the log
                if strFileName == strRSSI_MTK:
                    parseRSSI_MTK(dictRSSI, strLogPath)
                    b_hasRSSI = True

                # parse SarLog_DynAnt.txt
                if strFileName == strWIFI:
                    parseWIFI(dictWIFI, strLogPath)
                    b_hasWIFI = True

            # if log not exists, append initial dict
            listRSSI.append(dictRSSI)
            listWIFI.append(dictWIFI)

            # if there is no target log file in the folder, parse only SN
            if not b_hasRSSI:
                printLog("[W][ParseLog] Cannot find %s and %s by SN: %s" % (strRSSI_Intel, strRSSI_MTK, strSN))
            if not b_hasWIFI:
                printLog("[W][ParseLog] Cannot find %s by SN: %s" % (strWIFI, strSN))

        printLog("[I][parseLog] ------- Finish Parsing Log -------")
    except Exception as e:
        printLog("[E][parseLog] Unexpected Error: " + str(e))
    return listRSSI, listWIFI

# parse line contains 'CODE NO="010001A0"' and 'CODE NO="01000120"'
def parseRSSI_Intel(dictRSSI, strRSSIPath):
    try:
        with open(strRSSIPath, encoding="big5") as RSSILog:    # big5 for windows
            content = RSSILog.readlines()
            for line in content:
                # if the first line is PASS or FAIL, save to dict
                if content.index(line) == 0 and ("PASS" in line or "FAIL" in line):
                    dictRSSI["Test Result"] = line.strip("\n")

                if 'CODE NO="010001A0"' in line:
                    # line return <CODE NO="010001A0" KEYNAME="RSSI" VALUE1="-65" VALUE2="-63" ...
                    list_Info = line.split(" ")

                    # get VALUE1 figure
                    strSpec = list_Info[3].split("=")[1].strip('/>"')
                    dictRSSI["Spec"] = eval(strSpec)

                    # get VALUE2 figure
                    strMain = list_Info[4].split("=")[1].strip('/>"')
                    dictRSSI["Main"] = eval(strMain)

                if 'CODE NO="01000120"' in line:
                    # line return <CODE NO="01000120" KEYNAME="RSSI" VALUE1="-65" VALUE2="-62" ...
                    list_Info = line.split(" ")

                    strAux = list_Info[4].split("=")[1].strip('/>"')
                    dictRSSI["Aux"] = eval(strAux)
    except Exception as e:
        printLog("[E][parseRSSI_Intel] Unexpected Error: " + str(e))

# parse line contains "== ANT1 (sub) ==" and "== ANT2 (main) =="
def parseRSSI_MTK(dictRSSI, strRSSIPath):
    try:
        with open(strRSSIPath, encoding="big5") as RSSILog:    # big5 for windows
            content = RSSILog.readlines()
            for line in content:

                if "== ANT1 (sub) ==" in line:
                    idx = content.index(line)       # get the line index of keyword
                    for line in content[idx:]:      # start iteration from the line "== ANT1 (sub) =="
                        if "Threshold" in line:
                            strThreshold = line.split(": ")[1].strip("\n")
                            dictRSSI["Spec"] = eval(strThreshold)
                        if "Average" in line:
                            strMain = line.split(": ")[1].strip("\n")
                            dictRSSI["Main"] = eval(strMain)
                            break                   # find only one keyword "Average"

                if "== ANT2 (main) ==" in line:
                    idx = content.index(line)       # get the line index of keyword
                    for line in content[idx:]:
                        if "Average" in line:
                            strAux = line.split(": ")[1].strip("\n")
                            dictRSSI["Aux"] = eval(strAux)
                        if "PASS" in line or "FAIL" in line:
                            strResult = line.strip("\n")
                            dictRSSI["Test Result"] = strResult
                            break
    except Exception as e:
        print("[E][parseRSSI_MTK] Unexpected Error: " + str(e))

# parse the last line in SarLog_DynAnt.txt, target = "Value Match" or "Value Not Match"
def parseWIFI(dictWIFI, strWIFIPath):
    try:
        with open(strWIFIPath, encoding="big5") as WIFILog:    # big5 for windows
            content = WIFILog.readlines()
            # the target is in the last line
            if "Value Match" in content[-1] or "Value Not Match" in content[-1]:
                strResult = content[-1].strip("\n")
                dictWIFI["Result"] = strResult
    except Exception as e:
        printLog("[E][parseWIFI] Unexpected Error: " + str(e))

#/====================================================================\#
#|              Functions of parsing log to excel                     |#
#\====================================================================/#

# save two xlsx, RSSI.xlsx, WiFiSARQUERY.xlsx
def log_to_excel(listRSSI, listWIFI, strOutputFolder):
    printLog("[I][log_to_excel] ------- Parsing Log to Excel -------")

    try:
        # covert list of dict to DataFrame in order
        df_logRSSI = pd.DataFrame(listRSSI, columns=listRSSI[0].keys())
        df_logWIFI = pd.DataFrame(listWIFI, columns=listWIFI[0].keys())

        # set df_logRSSI index
        df_logRSSI.index.name = "Item"
        df_logRSSI.index += 1                   # index start with 1
        df_logRSSI = df_logRSSI.reset_index()   # reset index as normal column for constructing sheet

        # list the arguments for createing Excel workbooks
        list_df = [df_logRSSI, df_logWIFI]
        list_sheetname = ["RSSI", "wifisarquery"]
        list_fname = ["RSSI.xlsx", "WiFiSARQUERY.xlsx"]

        # new Excel workbook and sheets
        for i in range(len(list_df)):

            printLog("[I][log_to_excel] New workbook for %s" % list_fname[i])
            wb = openpyxl.Workbook()    # 新增 Excel 活頁
            wb.remove(wb['Sheet'])      # remove the default sheet when start a workbook

            # init vars
            df = list_df[i]
            str_sheet = list_sheetname[i]
            str_fname = list_fname[i]

            # set up sheet by DataFrame
            newSheet(wb, str_sheet, df)
            styleSheet(wb[str_sheet])

            # plot for RSSI_Report.xlsx
            if i == 0:
                newLineChart(wb[str_sheet], df)
                #print(df)
            wb.save(os.path.join(strOutputFolder, str_fname))     # save the worksheet as excel file
            printLog("[I][log_to_excel] %s generated" % str_fname)

        printLog("[I][log_to_excel] ------- Parsed Log to Excel -------")
    except Exception as e:
        printLog("[E][log_to_excel] Unexpected Error: " + str(e))

# new worksheets by DataFrame
def newSheet(workbook, strSheetName, df):
    try:
        workbook.create_sheet(strSheetName)
        for row in dataframe_to_rows(df, index=False, header=True):
            workbook[strSheetName].append(row)

        printLog("[I][newSheet] Sheet: %s Created" % strSheetName)
    except Exception as e:
        printLog("[E][newSheet] Unexpected Error: " + str(e))

# set cell width
def styleSheet(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))

    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value + 3
    dims.clear()

    # Enumerate the cells in the second row
    for cell in ws["1:1"]:
        cell.font = Font(color="F8F8FF")    # ghostwhite
        cell.fill = PatternFill(fgColor="292421", fill_type="solid")

# draw a line chart with 3 lines for RSSI
def newLineChart(ws, df):
    try:
        chart = LineChart()
        chart.title = "UNIT PN_M"
        chart.style = 18

        data = Reference(ws, min_col=4, min_row=1, max_col=6, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)

        # set daigram size
        chart.height = 10   # default is 7.5
        chart.width = 20

        # set y-axis according to the max value from DataFrame(df_logRSSI)
        chart.y_axis.scaling.max = -30

        # style the lines
        line_Main = chart.series[0]
        line_Main.graphicalProperties.line.solidFill = "00AAAA"     # navyblue
        line_Main.graphicalProperties.line.width = 25000

        line_Aux = chart.series[1]
        line_Aux.graphicalProperties.line.solidFill = "F08080"      # lightpink
        line_Aux.graphicalProperties.line.width = 25000

        line_Spec = chart.series[2]
        line_Spec.graphicalProperties.line.solidFill = "3D9140"     # cobaltgreen
        line_Spec.graphicalProperties.line.width = 25000

        ws.add_chart(chart, "I1")
    except Exception as e:
        printLog("[E][newLineChart] Unexpected Error: " + str(e))


#/====================================================================\#
#|                        Functions of printing log                   |#
#\====================================================================/#

def getDateTimeFormat():
    strDateTime = "[%s]" % (time.strftime("%Y/%m/%d %H:%M:%S", time.localtime()))
    return strDateTime

def printLog(strPrintLine):
    global strUser
    strFileName = os.path.basename(__file__).split('.')[0]
    #fileLog = codecs.open("/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY/" + strFileName + "_" + strUser + ".log", 'a', "utf-8")
    fileLog = codecs.open(strFileName + "_" + strUser + ".log", 'a', "utf-8")
    print(strPrintLine)
    fileLog.write("%s%s\r\n" % (getDateTimeFormat(), strPrintLine))
    fileLog.close()


# compress all files under dir "strZipDir", saving to "strZipPath" as .zip
def zip_all_files(strZipDir, strZipPath):
    try:
        printLog("[I][zip_all_files] Compressing files")
        with zipfile.ZipFile(strZipPath, "w") as zf:
            for dirPath, dirNames, fileNames in os.walk(strZipDir):
                for file in fileNames:
                    # omit zip file
                    if "zip" not in file:
                        zf.write(os.path.join(dirPath, file), file)
    except Exception as e:
        printLog("[E][zip_all_files] Unexpected Error: " + str(e))


if __name__ == "__main__":
    global strUser
    global logPath
    global b_localDebug

    # set user, logPath, and debug mode
    if(len(sys.argv) <= 1):
        strUser = "WillyWJ_Chen"
        b_localDebug = True
        logPath = os.getcwd()
    else:
        strUser = sys.argv[1]
        b_localDebug = False
        logPath = "/home/sanchez/Desktop/RDTool/Automation_RSSI_WiFiSARQUERY"   # log path for "WebComunicate_%s.txt"

    strFileName = os.path.basename(__file__).split('.')[0]

    printLog("========== Start ==========")
    printLog("[I][main] Python " + sys.version)
    printLog("[I][main] %s.py %s" % (strFileName, g_strVersion))
    printLog("[I][main] Decteing User: %s\n" % strUser)

    try:
        timeStartTime = time.time()

        ARW = Automation_RSSI_WiFiSARQUERY(strUser, b_localDebug)
        ARW.start()

        nSpendTime = int(time.time() - timeStartTime)
        printLog("[I][main] Spend Time: %d s" % (nSpendTime))

        hostname = socket.gethostname()
        IPAddr = get_host_ip()
        strDate = "%s" % (time.strftime("%Y-%m-%d", time.localtime()))
        strTime = "%s" % (time.strftime("%H:%M:%S", time.localtime()))
        data = {'pcip': IPAddr, 'pcname': hostname, 'account': strUser, 'toolid': 'Prj2112_RF_RSSI_QUERY',
                'rundate': strDate, 'runtime': strTime, 'executiontime': nSpendTime,
                'note': 'version: %s' % (g_strVersion), 'cycle': 1, 'sbtn': 'Send'}
        response = requests.post('http://10.110.140.43/LYUR/logYourUsageRecord.php', data)

    except Exception as e:
        printLog("[E][main] Unexpected Error: " + str(e))
    printLog("========== End ==========")

##! python3
##==============================================================================
##    Copyright (c) 2021 COMPAL Electronic Inc. All rights reserved.
##    This program contains proprietary and confidential information.
##    All rights reserved except as may be permitted by prior written consent.
##
##    Compal STiD NPSD Test Program Release Notification.
##
##    ModuleName:
##            LTE.py (Log to Excel)
##
##    Abstract:
##            Parsing log info to a excel with 4 sheets.
##              1. Read log file: parse -> store (a list of dict)
##              2. Read the INI threshold data: store as dict
##              3. New excel workbook: by openpyxl
##              4. Set worksheet according to Step 1: by dict and DataFrame
##              5. Set condition formating for each sheet
##                 according to Step 2: by dict
##              6. Save the workbook to xlsx file
##
##    Author:
##            25-Oct-2021 Willy Chen
##
##    Revision History:
##            Rev 1.0.0.2 03-Nov-2021 Willy
##                    Modified parsing logic for logs
##            Rev 1.0.0.1 25-Oct-2021 Willy
##                    First create.
##==============================================================================
import re
import os
import sys
import pandas as pd
import codecs
import time
import configparser
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Fill, colors
from openpyxl.formatting.rule import CellIsRule

# [Main]
g_strVersion = "3.0.0.2"

#[ParseLogPath]
g_strLogDir = "./Log/Pass/"


g_listKey = ["Power_dBm_CH15", "Power_dBm_CH21", "Power_dBm_CH24", "Current_mA_CH15", "Current_mA_CH21", "Current_mA_CH24", "dBm_LNA_ON", "dBm_LNA_Off",
 "Current_mA_3G_CH9750", "Current_mA_3G_CH2787", "Current_mA_2G_CH124", "dBm_CH9750", "dBm_CH2787", "dBm_2G_CH124", "dBm_CH124"]

#/====================================================================\#
#|               Functions of parsing target logs                     |#
#\====================================================================/#

def parseLog(listSNLogs):
    printLog("[I][parseLog] ------- Start Parsing Log -------")

    listLTE, listZigbee = [], []
    strLTEName, strZigbeeName = "GFI20_RF_LTE.log", "GFI20_RF_Zigbee.log"

    try:
        for strSN in listSNLogs:
            dictLTE = {
                "SN" : strSN,
                "dBm_CH9750" : None,
                "dBm_CH2787" : None,
                "dBm_2G_CH124" : None,
                "Current_mA_3G_CH9750" : None,
                "Current_mA_3G_CH2787" : None,
                "Current_mA_2G_CH124" : None,
                "dBm_CH124" : None }

            dictZigbee = {
                "SN" : strSN,
                "Power_dBm_CH15" : None,
                "Power_dBm_CH21" : None,
                "Power_dBm_CH24" : None,
                "dBm_LNA_ON" : None,
                "dBm_LNA_Off" : None,
                "Current_mA_CH15" : None,
                "Current_mA_CH21" : None,
                "Current_mA_CH24" : None }

            b_hasLTE, b_hasZigbee = False, False            # flag for checking if the target log exists
            strSNLog = os.path.join(g_strLogDir, strSN)     # set abspath for SN logs

            # iterate through log files in a SN folder (second layer)
            for strLogName in os.listdir(strSNLog):
                strLogPath = os.path.join(strSNLog, strLogName)

                # check GFI20_RF_LTE.log exists. If not, flag = False and parse only SN.
                reMatch = re.fullmatch("^.*RF_LTE\.log", strLogName)
                if(reMatch != None):
                    parseLTE(dictLTE, strLogPath)
                    b_hasLTE = True


                # parse GFI20_RF_Zigbee.log files
                reMatch = re.fullmatch("^.*RF_Zigbee\.log", strLogName)
                if(reMatch != None):
                    parseZigbee(dictZigbee, strLogPath)
                    b_hasZigbee = True

            # if log not exists, append initial dict
            listLTE.append(dictLTE)
            listZigbee.append(dictZigbee)

            # if there is no target log file in the folder, parse only SN
            if not b_hasLTE:
                printLog("\n[W][ParseLog] Cannot find log: %s\n" % os.path.join(strSN, strLTEName))
            if not b_hasZigbee:
                printLog("\n[W][ParseLog] Cannot find log: %s\n" % os.path.join(strSN, strZigbeeName))

        printLog("[I][parseLog] ------- Finish Parsing Log -------")
    except Exception as e:
        printLog("[E][parseLog] Unexpected Error: " + str(e))
    return listLTE, listZigbee

def parseLTE(dictLTE, strLTEPath):
    printLog("[I][parseLTE] Parse LTE log: %s" % strLTEPath.replace(g_strLogDir, ""))

    try:
        re_power = "Power: [+-]?[0-9]+\.?[0-9]*"
        re_current = "Current: [+-]?[0-9]+\.?[0-9]* A"
        re_RX_RSSI = "Rx RSSI: [+-]?[0-9]+\.?[0-9]* dBm"

        listPostfix = [" \n", " A\n", " dBm\n"]

        with open(strLTEPath, encoding='big5') as log:  # big5 for windows
            content = log.readlines()
            for line in content:
                if re.search("-+ LTE_3G Freq 897.4 -+", line) != None:
                    idx = content.index(line)       # get the line index of keyword
                    tmp_content = content[idx:]     # cut the content by the keyword
                    get_log_value(tmp_content, dictLTE, re_power, g_listKey[11], listPostfix[0], 1, False)
                    get_log_value(tmp_content, dictLTE, re_current, g_listKey[8], listPostfix[1], 1000, False)

                if re.search("-+ LTE_3G Freq 1950 -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictLTE, re_power, g_listKey[12], listPostfix[0], 1, False)
                    get_log_value(tmp_content, dictLTE, re_current, g_listKey[9], listPostfix[1], 1000, False)

                if re.search("-+ LTE_2G Freq 914.8 -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictLTE, re_power, g_listKey[13], listPostfix[0], 1, False)
                    get_log_value(tmp_content, dictLTE, re_current, g_listKey[10], listPostfix[1], 1000, False)

                if re.search("-+ LTE_2G Freq 959.8 -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictLTE, re_RX_RSSI, g_listKey[14], listPostfix[2], 1, True)

    except Exception as e:
        printLog("[E][parseLTE] Unexpected Error: " + str(e))


def parseZigbee(dictZigbee, strZigBeePath):
    printLog("[I][parseZigbee] Parse Zigbee log: %s" % strZigBeePath.replace(g_strLogDir, ""))

    try:
        re_power = "Power: [+-]?[0-9]+\.?[0-9]* dBm"
        re_current = "Current: [+-]?[0-9]+\.?[0-9]* A"
        re_RX_RSSI = "Rx RSSI: [+-]?[0-9]+\.?[0-9]* dBm"

        listPostfix = ["dBm\n", " A\n", " dBm\n"]

        with open(strZigBeePath, encoding="big5") as Zigbee:    # big5 for windows
            content = Zigbee.readlines()
            for line in content:
                if re.search("-+ ZIGBEE_2450 Freq 2425 -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictZigbee, re_power, g_listKey[0], listPostfix[0], 1, False)
                    get_log_value(tmp_content, dictZigbee, re_current, g_listKey[3], listPostfix[1], 1000, False)

                if re.search("-+ ZIGBEE_2450 Freq 2455 -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictZigbee, re_power, g_listKey[1], listPostfix[0], 1, False)
                    get_log_value(tmp_content, dictZigbee, re_current, g_listKey[4], listPostfix[1], 1000, False)

                if re.search("-+ ZIGBEE_2450 Freq 2470 -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictZigbee, re_power, g_listKey[2], listPostfix[0], 1, False)
                    get_log_value(tmp_content, dictZigbee, re_current, g_listKey[5], listPostfix[1], 1000, False)

                if re.search("-+ LNA ON -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictZigbee, re_RX_RSSI, g_listKey[6], listPostfix[2], 1,  False)

                if re.search("-+ LNA OFF -+", line) != None:
                    idx = content.index(line)
                    tmp_content = content[idx:]
                    get_log_value(tmp_content, dictZigbee, re_RX_RSSI, g_listKey[7], listPostfix[2], 1, False)

    except Exception as e:
        printLog("[E][parseZigbee] Unexpected Error: " + str(e))

# strPostfix for striping string, nUnit for (power/current) unit transform
def get_log_value(cut_content, dictInfo, re_target, strKey, strPostfix, nUnit, b_getMulti):
    try:
        for line in cut_content:
            # search pattern like "Power: (int/float) dBm"
            if re.search(re_target, line) != None:
                # get the figure of the line like "Power: 8.817 dBm\n"
                fValue = eval(line.split(": ")[1].strip(strPostfix))
                dictInfo[strKey] = fValue * nUnit
                if not b_getMulti:
                    break;

    except Exception as e:
        printLog("[E][get_log_value] Unexpected Error: " + str(e))

# merge two list of dict to single list of dict
def mergeLogs(listLTE, listZigbee):
    try:
        printLog("[I][mergeLogs] ------- Merging two Log data -------")
        # listLTE and listZigbee both has same length
        listInfo = [None] * len(listLTE)
        for i in range (0, len(listLTE)):
            listLTE[i].update(listZigbee[i])    # merge two dict
            listInfo[i] = listLTE[i]
        printLog("[I][mergeLogs] ------- Merged two Log data -------")
        return listInfo
    except Exception as e:
        printLog("[E][mergeLogs] Unexpected Error: " + str(e))
        return None


#/====================================================================\#
#|              Functions of parsing log to excel                     |#
#\====================================================================/#

def log_to_excel(listInfo):
    printLog("[I][log_to_excel] ------- Parsing Log to Excel -------")

    dictThreshold = {}  # store INI threshold ata for setting conditional formating
    try:
        # ========== get the threshold data from INI ==========
        printLog("[I][log_to_excel] ----- INI reading -----")
        for key in g_listKey:
            dictThreshold[key] = readINI(key)
        printLog("[I][log_to_excel] ----- INI read -----")

        # ========== New Excel workbook and sheets ==========
        df_logInfo = pd.DataFrame(listInfo)     # listInfo -> list of dict
        listSheetName = ["Zigbee_Power_Current", "Zigbee_LAN", "LTE_Current", "LTE_dBm"]
        listCol = [g_listKey[:6], g_listKey[6:8], g_listKey[8:11], g_listKey[11:15]]    # columns for each sheet above

        wb = openpyxl.Workbook()    # 新增 Excel 活頁
        wb.remove(wb['Sheet'])      # remove the default sheet when start a workbook

        printLog("[I][log_to_excel] ----- Excel Sheet Creating -----")
        for i in range(0, len(listSheetName)):
            newSheet(wb, listSheetName[i], df_logInfo[["SN"] + listCol[i]])
        printLog("[I][log_to_excel] ----- Excel Sheet Created -----")

        # modify cell font-color according to thershold that parsed from INI
        set_threshold_to_excel(wb, dictThreshold)

        wb.save('LTE.xlsx')     # save the worksheet as excel file

        printLog("[I][log_to_excel] ------- Parsed Log to Excel -------")
    except Exception as e:
        printLog("[E][log_to_excel] Unexpected Error: " + str(e))

# read INI values one by one by giving keys, then store to var dictThreshold
def readINI(strKey):
        try:
            config = configparser.ConfigParser()
            config.read(g_strINIPath)
            strMethod = 'Method%s' % g_nMethodIndex

            strValue = config.get(strMethod, strKey)

            # search pattern like "+-(int/float),+-(int/float)"
            if re.fullmatch("[+-]?[0-9]+\.?[0-9]*,[+-]?[0-9]+\.?[0-9]*", strValue):
                printLog("[I][readINI] %s = %s" % (strKey, strValue))
                return strValue
            else:
                printLog("[W][readINI] Read %s Fail !!" % strKey)
                sys.exit("Read %s Fail !!" % strKey)
        except Exception as e:
            printLog("[E][readINI] Error: %s" % str(e))
            sys.exit("Error: %s" % str(e))

# new worksheets by DataFrame
def newSheet(workbook, strSheetName, df_SheetCol):
    try:
        workbook.create_sheet(strSheetName)
        for row in dataframe_to_rows(df_SheetCol, index=False, header=True):
            workbook[strSheetName].append(row)
        printLog("[I][newSheet] Sheet: %s Created" % strSheetName)
    except Exception as e:
        printLog("[E][newSheet] Unexpected Error: " + str(e))

# set conditional formating for sheets by dictionay containg thershold data
def set_threshold_to_excel(workbook, dictThreshold):
    try:
        printLog("[I][set_threshold_to_excel] ----- threshold setting -----")

        # iterate through every worksheet to set conditional formatting
        for ws in workbook.worksheets:
            printLog("[I][set_threshold_to_excel] setting worksheet: %s" % ws.title)

            # iterate from Col 2 since Col 1 is the Serial Number(SN)
            for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                strStart, strEnd = None, None       # set the test range for cell e.g. A1:A10
                istInterval = []                    # set the threshold range for the formula below

                # check the column is not empty, col[0] is column name
                if len(col) > 1:
                    strStart = col[1].coordinate    # set starting cell for thershold testing
                    strEnd = col[-1].coordinate     # set ending cell

                    # get the thershold and store as interval for the formula below
                    strThreshold = dictThreshold[col[0].value]      # get the test thershold by the column name(col[0])
                    listInterval = strThreshold.split(",")

                red_text = Font(color="9C0006")                 # font-color: RED
                range_string = "%s:%s" % (strStart, strEnd)     # the value would be like A1:A10
                ws.conditional_formatting.add(range_string,
                    CellIsRule(operator='notBetween', formula=listInterval, stopIfTrue=True, font=red_text))

        printLog("[I][set_threshold_to_excel] ----- threshold set -----")
    except Exception as e:
        printLog("[E][set_threshold_to_excel] Unexpected Error: " + str(e))


#/====================================================================\#
#|              Functions of printing log of LTE.py                   |#
#\====================================================================/#

def getDateTimeFormat():
    strDateTime = "[%s]" % (time.strftime("%Y/%m/%d %H:%M:%S", time.localtime()))
    return strDateTime

def printLog(strPrintLine):
    strFileName = os.path.basename(__file__).split('.')[0]
    fileLog = codecs.open(g_strFileName + ".log", 'a', "utf-8")
    print(strPrintLine)
    fileLog.write("%s%s\r\n" % (getDateTimeFormat(), strPrintLine))
    fileLog.close()


if __name__ == "__main__":
    global g_strFileName, g_strINIPath, g_nMethodIndex
    g_strFileName = os.path.basename(__file__).split('.')[0]
    g_strINIPath = os.path.join(os.getcwd(), g_strFileName + ".ini")
    g_nMethodIndex = 1

    printLog("========== Start ==========")
    printLog("[I][main] Python " + sys.version)
    printLog("[I][main] %s.py %s" % (g_strFileName, g_strVersion))

    # ------------ find the target file --------------
    try:
        # get directory names of TryingLog (first layer)
        listSNLogs = os.listdir(g_strLogDir)
        # iterate through log files in a SN folder (second layer)
        listLTE, listZigbee = parseLog(listSNLogs)
        # merge data from two different log files
        listInfo = mergeLogs(listLTE, listZigbee)
        # save log to excel with thershold setting
        log_to_excel(listInfo)

    except Exception as e:
        printLog("[E][main] Unexpected Error: " + str(e))
    printLog("========== End ==========")
